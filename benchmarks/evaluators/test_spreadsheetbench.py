#!/usr/bin/env python3
import argparse
import asyncio
import datetime as _dt
import glob as _glob
import json
import os
import re
import subprocess
import sys
import tempfile
import textwrap
import time
from pathlib import Path

import openpyxl

from provider import (
    async_chat_completion,
    openai_compatible_api_key,
    openai_compatible_provider_label,
    provider_model,
)

                                                         
                                                  
                                                         

MODEL_CONFIG_GENERIC = {
    "model_provider": "openai_compatible",
    "model_name": "qwen3.6-plus",
    "api_key_env": "LLM_API_KEY",
    "base_url": None,
}

MODEL_CONFIG_CLAUDE = {
    "model_provider": "openai_compatible",
    "model_name": "anthropic.claude-opus-4-20250514",
    "api_key_env": "ANTHROPIC_API_KEY",
    "base_url": os.environ.get("LLM_BASE_URL", ""),
}

MODEL_CONFIG_CLAUDE_CLI = {
    "model_provider": "claude_cli",
    "model_name": "claude-opus-4-6",
    "api_key_env": "ANTHROPIC_AUTH_TOKEN",
    "base_url": "",
}

MODEL_CONFIG_KIMI_K2_6 = {
    "model_provider": "openai_compatible",
    "model_name": "kimi-k2.6",
    "api_key_env": "LLM_API_KEY",
    "base_url": None,
}

MODEL_CONFIG_DEEPSEEK_V4_PRO = {
    "model_provider": "openai_compatible",
    "model_name": "deepseek-v4-pro",
    "api_key_env": "LLM_API_KEY",
    "max_tokens": 16384,
    "timeout": 120,
}

MODEL_CONFIG_QWEN37MAX = {
    "model_provider": "openai_compatible",
    "model_name": "qwen3.7-max",
    "api_key_env": "LLM_API_KEY",
    "base_url": None,
}

MODEL_CONFIG_CLAUDE_COW = {
    "model_provider": "anthropic_api",
    "model_name": "claude-opus-4-6",
    "api_key_env": "ANTHROPIC_API_KEY",
    "base_url": os.environ.get("ANTHROPIC_BASE_URL", "https://api.anthropic.com"),
    "api_key": os.environ.get("ANTHROPIC_API_KEY"),
}

MAX_COMPLETION_TOKENS = 8192

async def call_llm_api(messages: list, model_name: str, api_key: str) -> tuple:
    import httpx
    payload = {"model": model_name, "messages": messages, "temperature": 0.1,
               "max_tokens": MAX_COMPLETION_TOKENS}
    async with httpx.AsyncClient(timeout=180.0) as client:
        data = await async_chat_completion(client, payload, api_key=api_key)
        return data["choices"][0]["message"]["content"], data.get("usage", {})

async def call_anthropic_api(messages: list, model_name: str, api_key: str, base_url: str) -> str:
    """Call Claude via Anthropic Messages API (provider-router compatible)."""
    import httpx
    url = f"{base_url.rstrip('/')}/v1/messages"
                           
    system_prompt = ""
    actual_messages = []
    for msg in messages:
        if msg["role"] == "system":
            system_prompt += msg["content"] + "\n"
        else:
            actual_messages.append(msg)
    payload = {
        "model": model_name,
        "messages": actual_messages,
        "max_tokens": MAX_COMPLETION_TOKENS,
        "temperature": 0.1,
    }
    if system_prompt.strip():
        payload["system"] = system_prompt.strip()
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    async with httpx.AsyncClient(timeout=180.0) as client:
        resp = await client.post(url, json=payload, headers=headers)
        resp.raise_for_status()
        data = resp.json()
        return data["content"][0]["text"]

async def call_claude_cli(messages: list, model_name: str) -> str:
    """passed claude -p (Claude Code CLI pipe mode) callmodel.

     messages columntableconvert as  system prompt file + stdin user prompt  form.
    multi- turnsfor passedat  user prompt  in with full history to implement.
    """
                      
    system_prompt = ""
    conversation_messages = []
    for msg in messages:
        if msg["role"] == "system":
            system_prompt += msg["content"] + "\n"
        else:
            conversation_messages.append(msg)

                                              
                                  
    if len(conversation_messages) == 1:
        user_prompt = conversation_messages[0]["content"]
    else:
        parts = []
        for msg in conversation_messages:
            if msg["role"] == "user":
                parts.append(f"[User]\n{msg['content']}")
            elif msg["role"] == "assistant":
                parts.append(f"[Assistant]\n{msg['content']}")
        parts.append("\n[Instructions]\nPlease provide your response to the latest user message above.")
        user_prompt = "\n\n---\n\n".join(parts)

                           
    sys_file = tempfile.NamedTemporaryFile("w", suffix=".txt", delete=False, encoding="utf-8")
    sys_file.write(system_prompt)
    sys_file.close()

    try:
        cmd = [
            "claude", "-p",
            "--model", model_name,
            "--bare",
            "--tools", "",
            "--no-session-persistence",
            "--system-prompt-file", sys_file.name,
        ]
        # Preserve the caller's Claude Code login and optional provider settings.
        # In particular, do not inject an empty token or a repository-local proxy.
        env = os.environ.copy()
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdin=asyncio.subprocess.PIPE,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            env=env,
        )
        stdout, stderr = await asyncio.wait_for(
            proc.communicate(input=user_prompt.encode("utf-8")),
            timeout=300.0,
        )
        if proc.returncode != 0:
            err_msg = stderr.decode("utf-8", errors="replace").strip()
            raise RuntimeError(f"claude CLI failed (rc={proc.returncode}): {err_msg[:500]}")
        return stdout.decode("utf-8", errors="replace").strip()
    finally:
        try:
            os.unlink(sys_file.name)
        except OSError:
            pass

async def call_llm(messages: list, config: dict) -> tuple:
    if config["model_provider"] == "claude_cli":
        content = await call_claude_cli(messages, config["model_name"])
        return content, {}
                                                                         
    if config["model_provider"] == "openai_compatible":
        api_key = openai_compatible_api_key(config["api_key_env"])
        return await call_llm_api(messages, config["model_name"], api_key)
    elif config["model_provider"] == "anthropic_api":
        api_key = os.environ.get(config["api_key_env"], "")
        if not api_key:
            raise RuntimeError(f"environment variable {config['api_key_env']} is not set")
        content = await call_anthropic_api(messages, config["model_name"], api_key, config["base_url"])
        return content, {}
    raise ValueError(f"not  model_provider: {config['model_provider']}")

def _datetime_to_float(dt: _dt.datetime) -> float:
    excel_start_date = _dt.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0

def _transform_value(v):
    if isinstance(v, bool):
        return round(float(v), 2)
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, _dt.time):
        return str(v)[:-3]
    if isinstance(v, _dt.datetime):
        return round(_datetime_to_float(v), 0)
    if isinstance(v, str):
        try:
            return round(float(v), 2)
        except ValueError:
            return v
    return v

def _compare_cell_value(v1, v2) -> bool:
    v1 = _transform_value(v1)
    v2 = _transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) is not type(v2):
        return False
    return v1 == v2

def _col_num2name(n: int) -> str:
    name = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        name = chr(65 + r) + name
    return name

def _col_name2num(name: str) -> int:
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num

def _parse_range(range_str: str):
    range_str = range_str.strip()
    if ":" not in range_str:
                      
        sc = "".join(ch for ch in range_str if ch.isalpha())
        sr = "".join(ch for ch in range_str if ch.isdigit())
        if not sr:          
            raise ValueError(f"Invalid cell reference: {range_str}")
        return (_col_name2num(sc), int(sr)), (_col_name2num(sc), int(sr))
    
    start_cell, end_cell = range_str.split(":")
    sc = "".join(ch for ch in start_cell if ch.isalpha())
    sr = "".join(ch for ch in start_cell if ch.isdigit())
    ec = "".join(ch for ch in end_cell if ch.isalpha())
    er = "".join(ch for ch in end_cell if ch.isdigit())
    
                                              
    if not sr and not er:
        return (_col_name2num(sc), 1), (_col_name2num(ec), 1048576)
    
    if not sr or not er:          
        raise ValueError(f"Invalid range: {range_str}")
    return (_col_name2num(sc), int(sr)), (_col_name2num(ec), int(er))

def _generate_cell_names(range_str: str):
    if ":" not in range_str:
        return [range_str]
    (sc, sr), (ec, er) = _parse_range(range_str)
                              
    if er - sr > 1000:
        er = sr + 99           
    cols = [_col_num2name(i) for i in range(sc, ec + 1)]
    return [f"{c}{r}" for c in cols for r in range(sr, er + 1)]

def _cell_level_compare(wb_gt, wb_proc, sheet_name: str, cell_range: str):
    if sheet_name not in wb_proc.sheetnames:
        return False, f"worksheet not found: {sheet_name}"
    ws_gt = wb_gt[sheet_name]
    ws_proc = wb_proc[sheet_name]
    try:
        for cn in _generate_cell_names(cell_range):
            cg = ws_gt[cn]
            cp = ws_proc[cn]
                                     
            if hasattr(cg, 'value') and hasattr(cp, 'value'):
                if not _compare_cell_value(cg.value, cp.value):
                    return False, f"value@{sheet_name}!{cn}: gt={cg.value!r} pred={cp.value!r}"
            else:
                return False, f"invalid cell object@{sheet_name}!{cn}"
        return True, ""
    except Exception as e:
        return False, f"invalid range '{cell_range}': {e}"

def compare_workbooks(gt_file: str, proc_file: str, answer_position: str):
    """single  test-case ratiofor , official semantics. returned  (ok, msg). """
    if not os.path.exists(proc_file):
        return False, "file not exist"
    try:
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
    except Exception as e:                
        return False, f"load error: {e}"
    try:
        ok_all = True
        msg_first = ""
        for scr in (answer_position or "").split(","):
            scr = scr.strip()
            if not scr:
                continue
                                     
            if "!" in scr:
                sheet_name, cell_range = scr.split("!", 1)
                sheet_name = sheet_name.strip().strip("'\"")
            else:
                sheet_name = wb_gt.sheetnames[0]
                cell_range = scr
            cell_range = cell_range.strip().strip("'\"")
            ok, msg = _cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range)
            if not ok:
                ok_all = False
                if not msg_first:
                    msg_first = msg
        return ok_all, msg_first
    finally:
        wb_gt.close()
        wb_proc.close()

_RUNNER_TEMPLATE = textwrap.dedent(
    """
    import os, sys, traceback
    INPUT_PATH = {input_path!r}
    OUTPUT_PATH = {output_path!r}
    try:
    {user_code_indented}
    except Exception:
        traceback.print_exc()
        sys.exit(2)
    """
)

_PATH_ASSIGN_RE = re.compile(r'^\s*(INPUT_PATH|OUTPUT_PATH)\s*=\s*.+$', re.MULTILINE)

def _strip_path_assignments(code: str) -> str:
    return _PATH_ASSIGN_RE.sub("", code)

def run_generated_code(code: str, input_path: str, output_path: str, timeout: int = 120):
    """sub processesexec  linesgenerate code, inject INPUT_PATH/OUTPUT_PATH. returned  (ok, err). """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    if os.path.exists(output_path):
        try:
            os.unlink(output_path)
        except OSError:
            pass
    cleaned = _strip_path_assignments(code)
    indented = textwrap.indent(cleaned, "    ")
    script = _RUNNER_TEMPLATE.format(
        input_path=input_path, output_path=output_path, user_code_indented=indented,
    )
    with tempfile.NamedTemporaryFile("w", suffix=".py", delete=False, encoding="utf-8") as f:
        f.write(script)
        tmp = f.name
    try:
        proc = subprocess.run(
            [sys.executable, tmp],
            capture_output=True, text=True,
            timeout=timeout if timeout and timeout > 0 else None,
        )
        if proc.returncode != 0:
            return False, (proc.stdout + "\n" + proc.stderr).strip()
        if not os.path.exists(output_path):
            return False, "output file was not created"
        return True, ""
    except subprocess.TimeoutExpired:
        return False, f"timeout after {timeout}s"
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass

                                                         
                                         
                                                         

def _preview_workbook(path: str, max_rows: int = 5, max_cols: int = 20) -> str:
    wb = openpyxl.load_workbook(path, data_only=False)
    chunks: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        chunks.append(
            f"## Sheet: {sheet_name}  "
            f"(dim={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column})"
        )
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows),
                                max_col=min(ws.max_column, max_cols), values_only=False):
            cells = []
            for cell in row:
                v = cell.value
                if v is None:
                    cells.append(f"{cell.coordinate}=")
                else:
                    s = str(v)
                    if len(s) > 40:
                        s = s[:37] + "..."
                    cells.append(f"{cell.coordinate}={s}")
            chunks.append(" | ".join(cells))
        if ws.max_row > max_rows:
            chunks.append(f"... ({ws.max_row - max_rows} more rows)")
        chunks.append("")
    wb.close()
    return "\n".join(chunks)

def extract_code(text: str) -> str:
    if "```" not in text:
        return text.strip()
    start = text.find("```")
    nl = text.find("\n", start)
    end = text.find("```", nl + 1)
    if nl == -1 or end == -1:
        return text.strip()
    return text[nl + 1: end].strip()

                                                         
                                                 
                                                         

def _find_test_cases(task_dir: str):
    """returned  [(case_no, input_path, answer_path), ...]. compatiblemulti-ways of namingname. """
    cases = []
    inputs = sorted(_glob.glob(os.path.join(task_dir, "*_input.xlsx")))
    for ip in inputs:
        no = os.path.basename(ip).split("_", 1)[0]
        ap = ip.replace("_input.xlsx", "_answer.xlsx")
        if os.path.exists(ap):
            cases.append((no, ip, ap))
    inits = sorted(_glob.glob(os.path.join(task_dir, "*_init.xlsx")))
    for ip in inits:
        no = os.path.basename(ip).split("_", 1)[0]
        ap = ip.replace("_init.xlsx", "_golden.xlsx")
        if os.path.exists(ap):
            cases.append((no, ip, ap))
    if not cases:
        bare_init = os.path.join(task_dir, "initial.xlsx")
        bare_gold = os.path.join(task_dir, "golden.xlsx")
        if os.path.exists(bare_init) and os.path.exists(bare_gold):
            cases.append(("1", bare_init, bare_gold))
    return cases

                                                         
           
                                                         

CODEGEN_SYSTEM = (
    "You are an expert Python programmer specializing in spreadsheet manipulation. "
    "You will be given a user instruction together with a preview of an input .xlsx file. "
    "Your job is to write a single self-contained Python script that reads the input file at "
    "the path stored in the variable INPUT_PATH, performs the requested manipulation, and saves "
    "the result to OUTPUT_PATH. Use only the standard library, openpyxl, and pandas. Do not print "
    "anything. Do not use input(). Do not hardcode file paths. Return ONLY the Python code inside "
    "a single ```python ... ``` fenced block."
)

def _build_system(skill_content: str) -> str:
    base = CODEGEN_SYSTEM
    if skill_content.strip():
        base += f"\n\n## Skill\n{skill_content.strip()}"
    return base

def _build_user(instruction: str, input_xlsx: str, instruction_type: str = "",
                answer_position: str = "") -> str:
    try:
        preview = _preview_workbook(input_xlsx)
    except Exception as e:                
        preview = f"(failed to preview workbook: {e})"
    extra = ""
    if instruction_type:
        extra += f"\nInstruction type: {instruction_type}"
    if answer_position:
        extra += f"\nExpected answer position: {answer_position}"
    return (
        f"# Instruction\n{instruction}\n{extra}\n\n"
        f"# Input spreadsheet preview\n{preview}\n\n"
        "# Task\n"
        "Write a Python script that reads the workbook from the variable `INPUT_PATH`, "
        "applies the instruction, and writes the modified workbook to `OUTPUT_PATH`. "
        "Preserve all other cells unchanged. "
        "The preview may be truncated — do not hardcode row counts or assume the data ends at "
        "the last previewed row; iterate over all actual rows in the workbook instead. "
        "Return only a ```python``` code block."
    )

                                                         
                                         
                                                         

async def run_multi_codegen(instruction: str, input_xlsx: str, gen_output: str,
                            instruction_type: str, answer_position: str,
                            skill_content: str, max_turns: int, model_config: dict,
                            exec_timeout: int = 120):
    system = _build_system(skill_content)
    user = _build_user(instruction, input_xlsx, instruction_type, answer_position)
    messages = [{"role": "system", "content": system},
                {"role": "user", "content": user}]
    conversation = []
    code = ""
    n_turns = 0
    last_err = ""

    for turn in range(max_turns):
        n_turns = turn + 1
                             
        raw = ""
        usage = {}
        for retry in range(3):
            try:
                if retry > 0:
                    await asyncio.sleep(3 * retry)
                raw, usage = await call_llm(messages, model_config)
                break
            except Exception as e:                
                last_err = f"llm-call-failed: {e}"
                es = str(e)
                if "429" in es or "Throttling" in es or "rate" in es.lower():
                    await asyncio.sleep(10 * (retry + 1))
                raw = ""
                usage = {}
        code = extract_code(raw)
        conversation.append({"role": "assistant", "content": raw, "usage": usage})
        messages.append({"role": "assistant", "content": raw})

        if not code.strip():
            feedback = ("No Python code block was found in your response. "
                        "Please return a complete Python script inside a ```python``` block.")
            messages.append({"role": "user", "content": feedback})
            conversation.append({"role": "user", "content": feedback})
            continue

        ok, err = await asyncio.to_thread(
            run_generated_code, code, input_xlsx, gen_output, exec_timeout)
        if ok:
            last_err = ""
            break
        last_err = err
        feedback = (
            f"The code raised an error during execution:\n\n```\n{err[:3000]}\n```\n\n"
            "Please fix the code and return a complete corrected Python script "
            "inside a ```python``` block."
        )
        messages.append({"role": "user", "content": feedback})
        conversation.append({"role": "user", "content": feedback})

    return {"code": code, "n_turns": n_turns, "conversation": conversation, "last_err": last_err}

                                                         
                                                 
                                                         

async def _run_one_sample(item: dict, sample_idx: int, skill_content: str,
                          semaphore: asyncio.Semaphore, model_config: dict, run_dir: Path,
                          max_turns: int, exec_timeout: int, task_timeout: int,
                          cases: list) -> dict:
    """single  timesindependentsampling: one  timesfull codegen + multi- case ratiofor , returned single  timesresults. """
    async with semaphore:
        task_id = str(item.get("id"))
        instruction = item.get("instruction", "")
        instruction_type = item.get("instruction_type", "")
        answer_position = item.get("answer_position", "")
        work_dir = run_dir / "work" / task_id / f"s{sample_idx}"
        work_dir.mkdir(parents=True, exist_ok=True)

        res = {"sample": sample_idx, "soft": 0.0, "hard": 0, "n_turns": 0,
               "code": "", "last_err": "", "conversation": [], "case_details": [],
               "error": ""}

        first_input = cases[0][1]
        gen_output = str(work_dir / "gen_output.xlsx")
        try:
            gen = await asyncio.wait_for(
                run_multi_codegen(
                    instruction, first_input, gen_output, instruction_type, answer_position,
                    skill_content, max_turns, model_config, exec_timeout),
                timeout=task_timeout)
        except asyncio.TimeoutError:
            res["error"] = f"task-timeout after {task_timeout}s"
            res["last_err"] = "task-timeout"
            return res
        except Exception as e:                
            res["error"] = f"codegen-failed: {e}"
            return res

        code = gen["code"]
        res["code"] = code
        res["n_turns"] = gen["n_turns"]
        res["last_err"] = gen.get("last_err", "")
        res["conversation"] = gen.get("conversation", [])
        try:
            (work_dir / "solution.py").write_text(code or "", encoding="utf-8")
        except Exception:
            pass

        if not code.strip():
            res["error"] = "no-code: " + (gen.get("last_err") or "")
            return res

        n_cases = len(cases)
        n_pass = 0
        case_details = []
        for (no, ip, ap) in cases:
            case_out = str(work_dir / f"output_{no}.xlsx")
            ok_exec, err = await asyncio.to_thread(
                run_generated_code, code, ip, case_out, exec_timeout)
            if not ok_exec:
                case_details.append({"case": no, "pass": False, "phase": "exec", "msg": err[:300]})
                continue
            ok_cmp, msg = compare_workbooks(ap, case_out, answer_position)
            if ok_cmp:
                n_pass += 1
                case_details.append({"case": no, "pass": True, "phase": "ok", "msg": ""})
            else:
                case_details.append({"case": no, "pass": False, "phase": "compare", "msg": msg[:300]})

        res["soft"] = n_pass / n_cases if n_cases else 0.0
        res["hard"] = 1 if (n_cases > 0 and n_pass == n_cases) else 0
        res["case_details"] = case_details
        return res

async def evaluate_task(item: dict, skill_content: str, semaphore: asyncio.Semaphore,
                        model_config: dict, data_root: str, run_dir: Path,
                        max_turns: int, exec_timeout: int, index: int, total: int,
                        task_timeout: int = 600, n_samples: int = 1) -> dict:
    task_id = str(item.get("id"))
    instruction_type = item.get("instruction_type", "")
    answer_position = item.get("answer_position", "")
    spreadsheet_path = item.get("spreadsheet_path", "")
    task_dir = os.path.join(data_root, spreadsheet_path)
    traces_dir = run_dir / "traces"

    base = {
        "task_id": task_id, "instruction_type": instruction_type,
        "answer_position": answer_position, "n_cases": 0, "n_pass": 0,
        "soft": 0.0, "hard": 0, "correct": False, "n_turns": 0, "n_samples": n_samples,
    }

    cases = _find_test_cases(task_dir)
    if not cases:
        base["error"] = f"no test cases under {task_dir}"
        print(f"  [{index}/{total}] {task_id}:  no test cases")
        return base

                                           
    samples = await asyncio.gather(*[
        _run_one_sample(item, s, skill_content, semaphore, model_config, run_dir,
                        max_turns, exec_timeout, task_timeout, cases)
        for s in range(n_samples)
    ])

    valid = [s for s in samples if not s.get("error")]
    n_valid = len(valid)
    hard_list = [s["hard"] for s in valid]
    soft_list = [s["soft"] for s in valid]
    pass_count = sum(hard_list)
                                   
    hard_majority = 1 if (n_valid > 0 and pass_count * 2 > n_valid) else 0
    soft_mean = sum(soft_list) / n_valid if n_valid else 0.0
    hard_mean = pass_count / n_valid if n_valid else 0.0
                                
    stable = (n_valid > 0 and (pass_count == 0 or pass_count == n_valid))

    n_cases = len(cases)
                             
    rep = next((s for s in valid if s["hard"] == 1), None) or (valid[0] if valid else samples[0])
    rep_pass = sum(1 for c in rep.get("case_details", []) if c.get("pass"))

    base.update({
        "n_cases": n_cases, "n_pass": rep_pass,
        "soft": soft_mean, "hard": hard_majority, "correct": bool(hard_majority),
        "n_turns": rep.get("n_turns", 0),
        "n_valid": n_valid, "pass_count": pass_count,
        "hard_mean": round(hard_mean, 4), "soft_mean": round(soft_mean, 4),
        "stable": stable, "sample_hard": hard_list,
        "case_details": rep.get("case_details", []),
    })
    if n_valid == 0:
        base["error"] = samples[0].get("error", "all-samples-failed")

    icon = ""if hard_majority else ""
    flag = "" if stable else " ~noisy"
    print(f"  [{index}/{total}] {task_id}: hard*={hard_majority} ({pass_count}/{n_valid}) "
          f"soft_avg={soft_mean:.2f} {icon}{flag}")

    _dump_trace(traces_dir, task_id, item, rep, base.get("case_details", []), base, samples)
    return base

def _dump_trace(traces_dir: Path, task_id: str, item: dict, gen: dict,
                case_details: list, base: dict, samples: list = None) -> None:
    traces_dir.mkdir(parents=True, exist_ok=True)
    trace = {
        "task_id": task_id,
        "instruction": item.get("instruction", ""),
        "instruction_type": item.get("instruction_type", ""),
        "answer_position": item.get("answer_position", ""),
        "code": gen.get("code", ""),
        "n_turns": gen.get("n_turns", 0),
        "last_err": gen.get("last_err", ""),
        "conversation": gen.get("conversation", []),
        "case_details": case_details,
        "soft": base.get("soft", 0.0),
        "hard": base.get("hard", 0),
        "timestamp": _dt.datetime.now().isoformat(),
    }
                           
    for k in ("n_samples", "n_valid", "pass_count", "hard_mean", "soft_mean",
              "stable", "sample_hard"):
        if k in base:
            trace[k] = base[k]
    if samples is not None:
        trace["samples"] = [
            {"sample": s.get("sample"), "hard": s.get("hard"), "soft": s.get("soft"),
             "n_turns": s.get("n_turns"), "error": s.get("error", ""),
             "first_fail": next((c["msg"] for c in s.get("case_details", []) if not c["pass"]), "")}
            for s in samples
        ]
    try:
        with open(traces_dir / f"trace_{task_id}.json", "w", encoding="utf-8") as f:
            json.dump(trace, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

                                                         
       
                                                         

def _load_items(data_file: str) -> list:
    with open(data_file, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, list) else data.get("items", [])

async def run_evaluation(args, model_config=None):
    print("=" * 70)
    print("SpreadsheetBench multi-codegen evaluation")
    print("=" * 70)

    if model_config is None:
        model_config = ACTIVE_MODEL_CONFIG
    model_config = dict(model_config)
    model_config["model_name"] = provider_model(model_config["model_name"])
    provider_label = (
        openai_compatible_provider_label()
        if model_config["model_provider"] == "openai_compatible"
        else model_config["model_provider"]
    )
    print(f"\n model: {provider_label} / {model_config['model_name']}")

    skill_path = Path(args.skill)
    skill_file = skill_path / "SKILL.md"
    if not skill_file.exists():
        print(f"Skill file not found: {skill_file}")
        sys.exit(1)
    skill_content = skill_file.read_text(encoding="utf-8")
    if skill_content.startswith("---"):
        end_idx = skill_content.find("---", 3)
        if end_idx != -1:
            skill_content = skill_content[end_idx + 3:].strip()
    print(f"Skill: {skill_path}")

    items = _load_items(args.data)
    if getattr(args, "ids", ""):
        wanted = {s.strip() for s in args.ids.split(",") if s.strip()}
        items = [it for it in items if str(it.get("id")) in wanted]
        print(f"directed backtest ids: {sorted(wanted)} → hit {len(items)} tasks")
    if args.limit and args.limit > 0:
        items = items[:args.limit]
    print(f"data: {len(items)} tasks from {args.data}")
    print(f"data_root: {args.data_root}")
    print(f"max_turns: {args.max_turns} | single  case timeout: {args.exec_timeout}s | concurrency: {args.max_concurrent}")
    print(f"self-consistency: n_samples={args.n_samples} (majority-vote aggregation hard) ")

    timestamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_base = Path(args.output_base) if args.output_base else (
        Path(__file__).parent.parent / "evolved" / "spreadsheetbench-solver")
    if not output_base.is_absolute():
        output_base = Path(__file__).parent.parent / output_base
    run_dir = output_base / f"{args.split}_run_{timestamp}"
    (run_dir / "traces").mkdir(parents=True, exist_ok=True)
    print(f"output: {run_dir}")

    semaphore = asyncio.Semaphore(args.max_concurrent)
    start = time.time()
    tasks = [
        evaluate_task(it, skill_content, semaphore, model_config, args.data_root,
                      run_dir, args.max_turns, args.exec_timeout, i + 1, len(items),
                      args.task_timeout, args.n_samples)
        for i, it in enumerate(items)
    ]
    results = await asyncio.gather(*tasks)
    elapsed = time.time() - start
    print(f"\n  elapsed: {elapsed:.1f}s")

    total = len(results)
    err = sum(1 for r in results if r.get("error"))
    soft_sum = sum(r["soft"] for r in results)
    hard_sum = sum(r["hard"] for r in results)
    soft_mean = soft_sum / total * 100 if total else 0
    hard_mean = hard_sum / total * 100 if total else 0

    print(f"\n{'='*70}")
    print(f"evaluation results")
    print(f"{'='*70}")
    print(f"  total task: {total}, execution exception: {err}")
    print(f"  Soft (avg pass rate): {soft_mean:.2f}%")
    print(f"  Hard (full  case passed): {hard_mean:.2f}% ({hard_sum}/{total})")

                            
    if args.n_samples > 1:
        noisy = [r for r in results if r.get("n_valid", 0) > 0 and not r.get("stable", True)]
                                                                  
        hm_vals = [r["hard_mean"] for r in results if "hard_mean" in r]
        sm_vals = [r["soft_mean"] for r in results if "soft_mean" in r]
        hard_exp = sum(hm_vals) / len(hm_vals) * 100 if hm_vals else 0
        soft_exp = sum(sm_vals) / len(sm_vals) * 100 if sm_vals else 0
        print(f"  ── self-consistency ──")
        print(f"  Hard(majority vote): {hard_mean:.2f}%  |  Hard(expected/per-run average): {hard_exp:.2f}%")
        print(f"  Soft(expected/per-run average): {soft_exp:.2f}%")
        print(f"  unstable task(cross-sample flips): {len(noisy)}/{total}")
        if noisy:
            ids = ", ".join(f"{r['task_id']}({r['pass_count']}/{r['n_valid']})" for r in noisy[:20])
            print(f"    {ids}")

                           
    types = sorted(set(r.get("instruction_type", "") for r in results))
    for t in types:
        if not t:
            continue
        sub = [r for r in results if r.get("instruction_type") == t]
        if not sub:
            continue
        s_soft = sum(r["soft"] for r in sub) / len(sub) * 100
        s_hard = sum(r["hard"] for r in sub) / len(sub) * 100
        print(f"  [{t}] n={len(sub)} Soft={s_soft:.2f}% Hard={s_hard:.2f}%")

             
    wrong = [r for r in results if not r["correct"]]
    if wrong:
        print(f"\n failed task ({len(wrong)}  total, showing first  20):")
        for w in wrong[:20]:
            fr = ""
            for c in w.get("case_details", []):
                if not c["pass"]:
                    fr = f"[{c['phase']}] {c['msg']}"
                    break
            if not fr and w.get("error"):
                fr = w["error"]
            print(f"  {w['task_id']} ({w.get('instruction_type','')}): {fr[:80]}")

        
    result_file = run_dir / f"results_{args.split}_{timestamp}.jsonl"
    with open(result_file, "w", encoding="utf-8") as f:
        for r in results:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    summary = {
        "timestamp": _dt.datetime.now().isoformat(),
        "run_timestamp": timestamp,
        "model": f"{provider_label}/{model_config['model_name']}",
        "skill_path": str(skill_path),
        "data_path": str(args.data),
        "split": args.split,
        "max_turns": args.max_turns,
        "exec_timeout": args.exec_timeout,
        "total": total,
        "error_count": err,
        "soft": round(soft_mean, 2),
        "hard": round(hard_mean, 2),
        "hard_pass": hard_sum,
        "n_samples": args.n_samples,
        "hard_expected": round(sum(r.get("hard_mean", r["hard"]) for r in results) / total * 100, 2) if total else 0,
        "soft_expected": round(sum(r.get("soft_mean", r["soft"]) for r in results) / total * 100, 2) if total else 0,
        "noisy_count": sum(1 for r in results if r.get("n_valid", 0) > 0 and not r.get("stable", True)),
        "elapsed_seconds": round(elapsed, 1),
        "results_file": str(result_file),
        "run_dir": str(run_dir),
    }
    summary_file = run_dir / f"report_{args.split}_{timestamp}.json"
    with open(summary_file, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(f"\n results: {result_file}")
    print(f"report: {summary_file}")
    print(f"run directory: {run_dir}")
    print("=" * 70)
    return summary

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="SpreadsheetBench multi-codegen batchevaluation")
    parser.add_argument("--skill", "-s", required=True, help="Skill directorypath (with  SKILL.md) ")
    parser.add_argument("--split", default="val", help="split namename: train/val/test (default val) ")
    parser.add_argument("--data", "-d", default="", help="items.json path (defaultby  split derive) ")
    parser.add_argument("--data-root", default="data/spreadsheetbench_verified_400",
                        dest="data_root", help="spreadsheet/{id} at directory")
    parser.add_argument("--split-dir", default="data/spreadsheetbench_split",
                        dest="split_dir", help="split directory (with  train/val/test/items.json) ")
    parser.add_argument("--max-turns", type=int, default=6, dest="max_turns", help="codegen max  turns count")
    parser.add_argument("--exec-timeout", type=int, default=120, dest="exec_timeout", help="single  case exec  linestimeout(s)")
    parser.add_argument("--max-concurrent", "-c", type=int, default=4, dest="max_concurrent", help="concurrency")
    parser.add_argument("--limit", type=int, default=0, help="only evaluationtop  N  task (0=full ) ")
    parser.add_argument("--ids", default="", help="only evaluationspecified  task id (comma separated) , use at directed backtest")
    parser.add_argument("--task-timeout", type=int, default=600, dest="task_timeout",
                        help="single  task overalltimeout(s), timeoutnote  as failed, avoidsingle  itemshang and blockfull ")
    parser.add_argument("--n-samples", "-n", type=int, default=1, dest="n_samples",
                        help="self-consistency sampling times count (per  task run independently N  times, majority-vote aggregation hard, default 1) ")
    parser.add_argument("--output-base", "-o", default="", dest="output_base", help="resultsoutputdirectory")
    parser.add_argument("--model", "-m", default="openai_compatible",
                        choices=["openai_compatible", "claude", "claude-cli", "claude-provider", "kimi-k2.6", "qwen3.7-max", "deepseek-v4-pro"],
                        help="modelchoose: openai_compatible(default), claude(HTTP), claude-cli(CLI pipe), claude-provider(provider-router), kimi-k2.6, qwen3.7-max")
    args = parser.parse_args()

    project_root = Path(__file__).parent.parent
    if not args.data:
        args.data = str(project_root / args.split_dir / args.split / "items.json")

                                   
    if args.model == "openai_compatible":
        if not Path(args.data).is_absolute():
            args.data = str(project_root / args.data)
        if not Path(args.data_root).is_absolute():
            args.data_root = str(project_root / args.data_root)
        if not Path(args.skill).is_absolute():
            args.skill = str(project_root / args.skill)

        if not Path(args.data).exists():
            print(f"dataset not found: {args.data}")
            sys.exit(1)
        if not Path(args.skill).exists():
            print(f"Skill directory not found: {args.skill}")
            sys.exit(1)
    
                         
    if args.model == "claude-cli":
        _active_config = MODEL_CONFIG_CLAUDE_CLI
    elif args.model == "claude":
        _active_config = MODEL_CONFIG_CLAUDE
    elif args.model == "claude-provider":
        _active_config = MODEL_CONFIG_CLAUDE_COW
    elif args.model == "kimi-k2.6":
        _active_config = MODEL_CONFIG_KIMI_K2_6
    elif args.model == "qwen3.7-max":
        _active_config = MODEL_CONFIG_QWEN37MAX
    elif args.model == "deepseek-v4-pro":
        _active_config = MODEL_CONFIG_DEEPSEEK_V4_PRO
    else:
        _active_config = MODEL_CONFIG_GENERIC
    
    asyncio.run(run_evaluation(args, _active_config))
